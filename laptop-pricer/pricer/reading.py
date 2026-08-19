"""Reading whatever file you actually have, rather than the one we wish you had.

Real exports arrive with report banners above the header, blank spacer rows,
currency symbols, thousands separators, trailing total rows, four different
date conventions and a stray byte-order mark. None of that should require you
to clean the file first.
"""
from __future__ import annotations

import csv
import datetime as dt
import re
from pathlib import Path

CURRENCY_SYMBOLS = {"£": "GBP", "$": "USD", "€": "EUR", "¥": "JPY", "zł": "PLN"}

DATE_FORMATS = [
    "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%d.%m.%Y", "%Y/%m/%d",
    "%d %b %Y", "%d %B %Y", "%b %d, %Y", "%B %d, %Y", "%d-%b-%Y", "%d-%b-%y",
    "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M", "%m/%d/%Y %H:%M", "%Y-%m-%dT%H:%M:%S",
    "%d/%m/%y", "%m/%d/%y", "%Y%m%d",
]

_NUM_CLEAN = re.compile(r"[^\d,.\-()]")
_EXCEL_EPOCH = dt.date(1899, 12, 30)     # Excel's off-by-two serial origin


def to_number(value) -> float | None:
    """Parse a price however it was written: £1,234.56  1.234,56  (99.00)  1 234."""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None

    negative = text.startswith("(") and text.endswith(")")
    text = _NUM_CLEAN.sub("", text).replace("(", "").replace(")", "")
    if not text or text in {"-", ".", ","}:
        return None

    if "," in text and "." in text:
        # whichever separator comes last is the decimal point
        text = (text.replace(",", "") if text.rfind(".") > text.rfind(",")
                else text.replace(".", "").replace(",", "."))
    elif "," in text:
        parts = text.split(",")
        # 1,234 is thousands; 1,50 is a decimal comma
        text = "".join(parts) if len(parts[-1]) == 3 and len(parts) > 1 else text.replace(",", ".")

    try:
        number = float(text)
    except ValueError:
        return None
    return -number if negative else number


def to_date(value, preferred: str | None = None) -> dt.date | None:
    """Parse a date from a string, a datetime, or an Excel serial number."""
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, (int, float)) and 20000 < value < 60000:
        return _EXCEL_EPOCH + dt.timedelta(days=int(value))

    text = str(value).strip()
    if not text:
        return None
    if re.fullmatch(r"\d{5}", text):                      # Excel serial as text
        return _EXCEL_EPOCH + dt.timedelta(days=int(text))

    formats = ([preferred] if preferred else []) + DATE_FORMATS
    for fmt in formats:
        try:
            return dt.datetime.strptime(text, fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def detect_currency(values: list) -> str | None:
    for value in values:
        for symbol, code in CURRENCY_SYMBOLS.items():
            if symbol in str(value):
                return code
    return None


def detect_date_format(values: list) -> str | None:
    """Pick the format that parses the most values, preferring day-first when
    both readings are valid - ambiguity here silently corrupts a time series."""
    best, best_hits = None, 0
    for fmt in DATE_FORMATS:
        hits = 0
        for value in values:
            try:
                dt.datetime.strptime(str(value).strip(), fmt)
                hits += 1
            except (ValueError, TypeError):
                pass
        if hits > best_hits:
            best, best_hits = fmt, hits
    return best if best_hits else None


# ---------- reading a whole file ----------

def _read_grid(path: Path, sheet: str | None = None) -> list[list]:
    """Every cell as a rectangular grid, before any header is assumed.

    Workbooks often open on a cover or summary sheet, so without an explicit
    sheet name the one carrying the most data wins rather than the first one.
    """
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xltx"}:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            names = [sheet] if sheet and sheet in wb.sheetnames else wb.sheetnames
            best, best_cells = [], -1
            for name in names:
                grid = [["" if c is None else c for c in row]
                        for row in wb[name].iter_rows(values_only=True)]
                cells = sum(1 for row in grid for c in row if str(c).strip())
                if cells > best_cells:
                    best, best_cells = grid, cells
            return best
        finally:
            wb.close()

    with open(path, newline="", encoding="utf-8-sig", errors="replace") as fh:
        sample = fh.read(16384)
        fh.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        return [row for row in csv.reader(fh, dialect)]


def _looks_like_header(row: list, following: list[list]) -> float:
    """Score a row on how much it behaves like a header rather than data."""
    cells = [str(c).strip() for c in row]
    filled = [c for c in cells if c]
    if len(filled) < 2:
        return 0.0

    score = 0.0
    score += min(len(filled) / 6.0, 1.0)                                  # reasonably wide
    if len(set(filled)) == len(filled):
        score += 0.6                                                       # unique labels
    texty = sum(1 for c in filled if to_number(c) is None and to_date(c) is None)
    score += 1.2 * texty / len(filled)                                     # headers are words
    score += 0.5 * sum(1 for c in filled if len(c) <= 30) / len(filled)    # and short ones

    # a header is followed by rows that do contain numbers or dates
    below = [c for row in following[:5] for c in row if str(c).strip()]
    if below:
        typed = sum(1 for c in below if to_number(c) is not None or to_date(c) is not None)
        score += 1.0 * typed / len(below)
    return score


def find_header_row(grid: list[list], search_depth: int = 25) -> int:
    best_index, best_score = 0, -1.0
    for i, row in enumerate(grid[:search_depth]):
        score = _looks_like_header(row, grid[i + 1:])
        if score > best_score:
            best_index, best_score = i, score
    return best_index


def read_rows(path: str | Path, header_row: int | None = None,
              sheet: str | None = None) -> tuple[list[dict], dict]:
    """Return (rows, metadata). Junk above the header, blank rows and trailing
    total rows are dropped rather than becoming phantom observations."""
    path = Path(path)
    grid = _read_grid(path, sheet)
    if not grid:
        return [], {"path": str(path), "rows": 0, "note": "file is empty"}

    header_index = find_header_row(grid) if header_row is None else header_row
    header = [str(c).strip() for c in grid[header_index]]
    header = [h or f"column_{i+1}" for i, h in enumerate(header)]

    rows, skipped_blank, skipped_total = [], 0, 0
    for raw in grid[header_index + 1:]:
        cells = list(raw) + [""] * (len(header) - len(raw))
        values = [str(c).strip() for c in cells[:len(header)]]
        if not any(values):
            skipped_blank += 1
            continue
        # a summary line: some cell is literally the word "total". A genuine
        # item description is never exactly that, so this is safe to drop.
        if any(re.fullmatch(r"(?i)(grand ?total|sub ?total|totals?|sum)s?\s*:?", v) for v in values):
            skipped_total += 1
            continue
        rows.append(dict(zip(header, cells[:len(header)])))

    return rows, {
        "path": str(path), "format": path.suffix.lower().lstrip(".") or "csv",
        "header_row": header_index, "columns": header, "rows": len(rows),
        "skipped_blank": skipped_blank, "skipped_total_rows": skipped_total,
        "junk_rows_above_header": header_index,
    }


def compose_title(row: dict, spec) -> str:
    """Build a description from one column, or from several.

    Plenty of stock systems never store a title at all - brand, model, CPU and
    memory each live in their own column. `raw_title` may therefore be a list,
    and the pieces are joined into the sentence the parser expects.
    """
    if not isinstance(spec, (list, tuple)):
        return str(row.get(spec, "") or "").strip()

    parts = []
    for column in spec:
        value = str(row.get(column, "") or "").strip()
        if not value:
            continue
        # a bare "16" in a memory column means 16GB
        if re.fullmatch(r"\d{1,4}", value) and re.search(r"(?i)ram|memor|storage|ssd|hdd|disk|drive", column):
            value += "GB"
        parts.append(value)
    return " ".join(parts)
